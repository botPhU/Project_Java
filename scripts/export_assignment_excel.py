# -*- coding: utf-8 -*-
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


OUTPUT_PATH = Path(__file__).resolve().parents[1] / "PHAN_CONG_CHI_TIET_TEAM.xlsx"


SUMMARY_ROWS = [
    ["Bạn - Trưởng nhóm", "Frontend toàn bộ và chốt tích hợp", "ui", "Đã có đủ file frontend để tiếp tục làm việc"],
    ["Bảo Huy", "Lead backend, trực tiếp code phần lõi backend", "backend/trend-dashboard; backend/datasource-sync; review toàn bộ backend", "Đã có file sẵn + đã tạo đủ file khung backend lõi"],
    ["Trần Lương Minh", "Backend auth, user, paper", "backend/auth-user; backend/paper-search", "Đã có file sẵn + đã tạo đủ file khung cần tiếp tục"],
    ["Trương Ngọc Duy", "Backend bookmark, notification, admin", "backend/bookmark-follow; backend/notification; backend/admin-system", "Đã có file sẵn + đã tạo đủ file khung cần tiếp tục"],
    ["Nguyễn Chí Hải", "Code admin support, sync support, release support", "backend/admin-system; phối hợp backend/datasource-sync", "Đã có file để code trực tiếp"],
    ["Trần Hà Đức Huỳnh", "Code bookmark, notification, business flow support", "backend/bookmark-follow; backend/notification; phối hợp backend/paper-search", "Đã có file để code trực tiếp"],
]


BRANCH_ROWS = [
    ["Bạn - Trưởng nhóm", "ui", "Frontend", "Làm toàn bộ giao diện và kết nối API frontend"],
    ["Bảo Huy", "backend/trend-dashboard", "Trend / Dashboard", "Lead backend và trực tiếp làm xu hướng theo keyword, dashboard tổng quan"],
    ["Bảo Huy", "backend/datasource-sync", "Datasource / Sync", "Lead backend và trực tiếp làm tích hợp OpenAlex, Crossref, Semantic Scholar, đồng bộ dữ liệu"],
    ["Bảo Huy", "Review toàn backend", "Common / Database", "Chốt cấu trúc dùng chung, review entity, config, exception, merge backend"],
    ["Trần Lương Minh", "backend/auth-user", "Auth / User", "Đăng ký, đăng nhập, users/me"],
    ["Trần Lương Minh", "backend/paper-search", "Paper", "Tìm kiếm bài báo, chi tiết bài báo, phân trang"],
    ["Trương Ngọc Duy", "backend/bookmark-follow", "Bookmark / Follow", "Lưu bài báo, follow topic hoặc keyword"],
    ["Trương Ngọc Duy", "backend/notification", "Notification", "Danh sách thông báo, đánh dấu đã đọc"],
    ["Trương Ngọc Duy", "backend/admin-system", "Admin", "Quản trị user, sync trigger, chức năng admin"],
    ["Nguyễn Chí Hải", "backend/admin-system", "Admin / Sync support", "Trạng thái hệ thống, hỗ trợ sync run, hỗ trợ release backend"],
    ["Nguyễn Chí Hải", "phối hợp backend/datasource-sync", "Datasource support", "Hỗ trợ cấu hình và bàn giao kỹ thuật cho sync"],
    ["Trần Hà Đức Huỳnh", "backend/bookmark-follow", "Bookmark", "Hoàn thiện flow bookmark và follow theo nghiệp vụ"],
    ["Trần Hà Đức Huỳnh", "backend/notification", "Notification", "Hoàn thiện notification và business wording"],
    ["Trần Hà Đức Huỳnh", "phối hợp backend/paper-search", "Business flow", "Hỗ trợ flow người dùng giữa search, bookmark, notification"],
]


BACKEND_ROWS = [
    [
        "Bảo Huy",
        "backend/trend-dashboard; backend/datasource-sync; review toàn backend",
        "common; trend; dashboard; datasource; sync",
        "backend/src/main/java/com/swp/scijournal/common/model/BaseEntity.java\nbackend/src/main/java/com/swp/scijournal/common/web/ApiResponse.java\nbackend/src/main/java/com/swp/scijournal/common/config/SecurityConfig.java\nbackend/src/main/resources/application.yml\nbackend/src/main/java/com/swp/scijournal/trend/controller/TrendController.java\nbackend/src/main/java/com/swp/scijournal/sync/job/AcademicSyncJob.java",
        "backend/src/main/java/com/swp/scijournal/common/exception/GlobalExceptionHandler.java\nbackend/src/main/java/com/swp/scijournal/common/exception/BusinessException.java\nbackend/src/main/java/com/swp/scijournal/common/config/JacksonConfig.java\nbackend/src/main/java/com/swp/scijournal/common/config/OpenApiDescription.md\nbackend/src/main/java/com/swp/scijournal/trend/entity/PublicationTrend.java\nbackend/src/main/java/com/swp/scijournal/trend/repository/PublicationTrendRepository.java\nbackend/src/main/java/com/swp/scijournal/trend/dto/TrendPointResponse.java\nbackend/src/main/java/com/swp/scijournal/trend/dto/KeywordTrendResponse.java\nbackend/src/main/java/com/swp/scijournal/trend/service/TrendService.java\nbackend/src/main/java/com/swp/scijournal/dashboard/controller/DashboardController.java\nbackend/src/main/java/com/swp/scijournal/dashboard/dto/DashboardOverviewResponse.java\nbackend/src/main/java/com/swp/scijournal/dashboard/service/DashboardService.java\nbackend/src/main/java/com/swp/scijournal/datasource/openalex/OpenAlexClient.java\nbackend/src/main/java/com/swp/scijournal/datasource/crossref/CrossrefClient.java\nbackend/src/main/java/com/swp/scijournal/datasource/semanticscholar/SemanticScholarClient.java\nbackend/src/main/java/com/swp/scijournal/datasource/dto/ExternalPaperRecord.java\nbackend/src/main/java/com/swp/scijournal/datasource/entity/ApiDataSource.java\nbackend/src/main/java/com/swp/scijournal/datasource/repository/ApiDataSourceRepository.java\nbackend/src/main/java/com/swp/scijournal/sync/service/SyncService.java",
        "Lead backend\nChuẩn hóa exception và config dùng chung\nLàm trend theo keyword và topic\nLàm dashboard overview\nLàm tích hợp academic source\nLàm sync thủ công và định kỳ\nReview backend trước khi merge",
        "GET /api/v1/trends/keywords/{keyword}\nGET /api/v1/dashboard/overview\nChuẩn response lỗi backend\nSync nguồn ngoài",
        "Đã có đủ file khung để bắt đầu làm"
    ],
    [
        "Trần Lương Minh",
        "backend/auth-user; backend/paper-search",
        "auth; user; paper",
        "backend/src/main/java/com/swp/scijournal/auth/controller/AuthController.java\nbackend/src/main/java/com/swp/scijournal/auth/service/AuthService.java\nbackend/src/main/java/com/swp/scijournal/auth/service/DatabaseUserDetailsService.java\nbackend/src/main/java/com/swp/scijournal/user/entity/User.java\nbackend/src/main/java/com/swp/scijournal/user/repository/UserRepository.java\nbackend/src/main/java/com/swp/scijournal/paper/controller/PaperController.java\nbackend/src/main/java/com/swp/scijournal/paper/service/PaperService.java\nbackend/src/main/java/com/swp/scijournal/paper/repository/ResearchPaperRepository.java",
        "backend/src/main/java/com/swp/scijournal/user/controller/UserController.java\nbackend/src/main/java/com/swp/scijournal/user/dto/UserProfileResponse.java\nbackend/src/main/java/com/swp/scijournal/user/service/UserService.java\nbackend/src/main/java/com/swp/scijournal/paper/dto/PaperSearchRequest.java\nbackend/src/main/java/com/swp/scijournal/paper/dto/PaperSearchResponse.java\nbackend/src/main/java/com/swp/scijournal/paper/service/PaperQueryService.java",
        "Hoàn thiện đăng ký và đăng nhập\nHoàn thiện users/me\nPhối hợp với Bảo Huy để chuẩn hóa lỗi auth\nThêm phân trang và validate tìm kiếm paper\nỔn định paper detail",
        "POST /api/v1/auth/register\nPOST /api/v1/auth/login\nGET /api/v1/users/me\nGET /api/v1/papers\nGET /api/v1/papers/{id}",
        "Đã có đủ file khung để bắt đầu làm"
    ],
    [
        "Trương Ngọc Duy",
        "backend/bookmark-follow; backend/notification; backend/admin-system",
        "bookmark; notification; admin",
        "backend/src/main/java/com/swp/scijournal/admin/package-info.java\nbackend/src/main/java/com/swp/scijournal/bookmark/package-info.java\nbackend/src/main/java/com/swp/scijournal/notification/package-info.java",
        "backend/src/main/java/com/swp/scijournal/bookmark/controller/BookmarkController.java\nbackend/src/main/java/com/swp/scijournal/bookmark/entity/Bookmark.java\nbackend/src/main/java/com/swp/scijournal/bookmark/entity/FollowTarget.java\nbackend/src/main/java/com/swp/scijournal/bookmark/repository/BookmarkRepository.java\nbackend/src/main/java/com/swp/scijournal/bookmark/repository/FollowTargetRepository.java\nbackend/src/main/java/com/swp/scijournal/bookmark/dto/BookmarkResponse.java\nbackend/src/main/java/com/swp/scijournal/bookmark/service/BookmarkService.java\nbackend/src/main/java/com/swp/scijournal/notification/controller/NotificationController.java\nbackend/src/main/java/com/swp/scijournal/notification/entity/Notification.java\nbackend/src/main/java/com/swp/scijournal/notification/repository/NotificationRepository.java\nbackend/src/main/java/com/swp/scijournal/notification/dto/NotificationResponse.java\nbackend/src/main/java/com/swp/scijournal/notification/service/NotificationService.java\nbackend/src/main/java/com/swp/scijournal/admin/controller/AdminController.java\nbackend/src/main/java/com/swp/scijournal/admin/dto/AdminUserResponse.java\nbackend/src/main/java/com/swp/scijournal/admin/dto/SyncRunResponse.java\nbackend/src/main/java/com/swp/scijournal/admin/service/AdminService.java",
        "Làm bookmark và follow\nLàm notification\nLàm chức năng admin\nPhối hợp với Bảo Huy khi đụng config hoặc sync",
        "POST /api/v1/bookmarks/papers/{paperId}\nPOST /api/v1/follows/topics/{topicId}\nGET /api/v1/notifications\nPATCH /api/v1/notifications/{id}/read\nPOST /api/v1/admin/sync/run",
        "Đã có đủ file khung để bắt đầu làm"
    ],
    [
        "Nguyễn Chí Hải",
        "backend/admin-system; phối hợp backend/datasource-sync",
        "admin; sync support; release support",
        "backend/src/main/java/com/swp/scijournal/admin/controller/AdminController.java\nbackend/src/main/java/com/swp/scijournal/admin/dto/AdminUserResponse.java\nbackend/src/main/java/com/swp/scijournal/admin/dto/SyncRunResponse.java\nbackend/src/main/java/com/swp/scijournal/admin/service/AdminService.java\nbackend/src/main/java/com/swp/scijournal/sync/service/SyncService.java\nbackend/src/main/resources/application.yml\nREADME.md",
        "backend/src/main/java/com/swp/scijournal/admin/dto/SystemHealthResponse.java\nbackend/src/main/java/com/swp/scijournal/admin/dto/SystemConfigResponse.java\nbackend/src/main/java/com/swp/scijournal/admin/service/SystemMonitorService.java\nbackend/src/main/java/com/swp/scijournal/common/config/ReleaseNote.md",
        "Hoàn thiện API chạy sync thủ công\nLàm response trạng thái hệ thống cho admin\nLàm phần hỗ trợ kiểm tra cấu hình backend\nChuẩn bị tài liệu release và hướng dẫn cài đặt",
        "POST /api/v1/admin/sync/run\nDữ liệu trạng thái hệ thống cho admin\nInstall guide\nRelease note",
        "Đã có file để code trực tiếp"
    ],
    [
        "Trần Hà Đức Huỳnh",
        "backend/bookmark-follow; backend/notification; phối hợp backend/paper-search",
        "bookmark; notification; business flow support",
        "backend/src/main/java/com/swp/scijournal/bookmark/controller/BookmarkController.java\nbackend/src/main/java/com/swp/scijournal/bookmark/entity/Bookmark.java\nbackend/src/main/java/com/swp/scijournal/bookmark/entity/FollowTarget.java\nbackend/src/main/java/com/swp/scijournal/bookmark/repository/BookmarkRepository.java\nbackend/src/main/java/com/swp/scijournal/bookmark/repository/FollowTargetRepository.java\nbackend/src/main/java/com/swp/scijournal/bookmark/dto/BookmarkResponse.java\nbackend/src/main/java/com/swp/scijournal/bookmark/service/BookmarkService.java\nbackend/src/main/java/com/swp/scijournal/notification/controller/NotificationController.java\nbackend/src/main/java/com/swp/scijournal/notification/entity/Notification.java\nbackend/src/main/java/com/swp/scijournal/notification/repository/NotificationRepository.java\nbackend/src/main/java/com/swp/scijournal/notification/dto/NotificationResponse.java\nbackend/src/main/java/com/swp/scijournal/notification/service/NotificationService.java",
        "backend/src/main/java/com/swp/scijournal/bookmark/dto/FollowTopicResponse.java\nbackend/src/main/java/com/swp/scijournal/notification/dto/NotificationSummaryResponse.java\nbackend/src/main/java/com/swp/scijournal/notification/service/NotificationRuleService.java",
        "Hoàn thiện follow keyword hoặc topic\nHoàn thiện rule tạo notification\nLàm wording response bám nghiệp vụ hơn\nPhối hợp tối ưu luồng người dùng giữa search, bookmark và notification",
        "POST /api/v1/follows/topics/{topicId}\nGET /api/v1/notifications\nPATCH /api/v1/notifications/{id}/read\nLogic rule notification",
        "Đã có file để code trực tiếp"
    ],
]


FRONTEND_ROWS = [
    ["Người phụ trách", "Bạn - Trưởng nhóm"],
    ["Nhánh làm việc", "ui"],
    ["Phạm vi", "Toàn bộ frontend"],
    ["Trạng thái", "Đã có đầy đủ file frontend để tiếp tục phát triển"],
    [
        "File frontend đang giữ",
        "frontend/src/App.jsx\nfrontend/src/main.jsx\nfrontend/src/components/AppLayout.jsx\nfrontend/src/components/Header.jsx\nfrontend/src/components/Footer.jsx\nfrontend/src/components/ProtectedRoute.jsx\nfrontend/src/components/PublicOnlyRoute.jsx\nfrontend/src/context/AuthContext.jsx\nfrontend/src/pages/LoginPage.jsx\nfrontend/src/pages/PaperSearchPage.jsx\nfrontend/src/pages/PaperDetailPage.jsx\nfrontend/src/pages/DashboardPage.jsx\nfrontend/src/pages/AdminPage.jsx\nfrontend/src/services/apiClient.js\nfrontend/src/services/authService.js\nfrontend/src/services/paperService.js\nfrontend/src/styles/global.css\nfrontend/vite.config.js"
    ],
]


HEADER_FILL = PatternFill("solid", fgColor="1F6F61")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E3E0"),
    right=Side(style="thin", color="D9E3E0"),
    top=Side(style="thin", color="D9E3E0"),
    bottom=Side(style="thin", color="D9E3E0"),
)


def style_sheet(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
    ws.freeze_panes = "A2"


def set_widths(ws, widths):
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def make_header(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_summary_sheet(wb):
    ws = wb.create_sheet("Tong quan")
    make_header(ws, ["Thành viên", "Vai trò", "Nhánh làm việc", "Trạng thái file khung"])
    for row in SUMMARY_ROWS:
        ws.append(row)
    style_sheet(ws)
    set_widths(ws, {"A": 24, "B": 42, "C": 58, "D": 38})


def build_branch_sheet(wb):
    ws = wb.create_sheet("Nhanh backend")
    make_header(ws, ["Thành viên", "Nhánh", "Module", "Mục tiêu làm việc"])
    for row in BRANCH_ROWS:
        ws.append(row)
    style_sheet(ws)
    set_widths(ws, {"A": 24, "B": 38, "C": 28, "D": 50})


def build_backend_sheet(wb):
    ws = wb.create_sheet("Backend chi tiet")
    make_header(
        ws,
        [
            "Thành viên",
            "Nhánh làm việc",
            "Module",
            "File hiện có",
            "File khung đã tạo thêm",
            "Việc cụ thể",
            "Deliverable",
            "Trạng thái",
        ],
    )
    for row in BACKEND_ROWS:
        ws.append(row)
    style_sheet(ws)
    set_widths(ws, {"A": 24, "B": 44, "C": 26, "D": 52, "E": 56, "F": 34, "G": 30, "H": 26})


def build_frontend_sheet(wb):
    ws = wb.create_sheet("Frontend cua ban")
    make_header(ws, ["Mục", "Nội dung"])
    for row in FRONTEND_ROWS:
        ws.append(row)
    style_sheet(ws)
    set_widths(ws, {"A": 24, "B": 96})


def build_note_sheet(wb):
    ws = wb.create_sheet("Ghi chu")
    ws.append(["Ghi chú quan trọng"])
    ws["A1"].font = Font(color="FFFFFF", bold=True)
    ws["A1"].fill = HEADER_FILL
    ws["A2"] = "Bảo Huy là lead backend và đồng thời trực tiếp code các phần lõi backend."
    ws["A3"] = "Nguyễn Chí Hải và Trần Hà Đức Huỳnh đều đã được chuyển sang có phần code backend thực tế."
    ws["A4"] = "Không còn ghi hai bạn này ở vai trò test chéo hoặc review chéo trong bảng phân công."
    style_sheet(ws)
    set_widths(ws, {"A": 120})


def main():
    wb = Workbook()
    wb.remove(wb.active)

    build_summary_sheet(wb)
    build_branch_sheet(wb)
    build_backend_sheet(wb)
    build_frontend_sheet(wb)
    build_note_sheet(wb)

    try:
        wb.save(OUTPUT_PATH)
        print(f"Da tao file Excel: {OUTPUT_PATH}")
    except PermissionError:
        fallback_path = OUTPUT_PATH.with_name("PHAN_CONG_CHI_TIET_TEAM_CAP_NHAT.xlsx")
        wb.save(fallback_path)
        print(f"Da tao file Excel: {fallback_path}")


if __name__ == "__main__":
    main()
